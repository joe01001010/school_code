import requests
import bs4
import pandas as pd
from openpyxl.styles import Border, Side
from openpyxl import load_workbook
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


def main():
    url2022 = "https://openaccess.thecvf.com/CVPR2022?day=all"
    url2023 = "https://openaccess.thecvf.com/CVPR2023?day=all"
    url2024 = "https://openaccess.thecvf.com/CVPR2024?day=all"
    output_file = 'most_contributions.xlsx'

    # These expect keyword arguments of the url to query and will return the unparsed response
    raw_response_2022 = get_most_contributions(url=url2022)
    raw_response_2023 = get_most_contributions(url=url2023)
    raw_response_2024 = get_most_contributions(url=url2024)

    # These functions expect the unparsed response
    # will return the author dictionaries with contributions as values
    authors_2022 = parse_authors(raw_html=raw_response_2022)
    authors_2023 = parse_authors(raw_html=raw_response_2023)
    authors_2024 = parse_authors(raw_html=raw_response_2024)

    # This function expects all the data and will return all the top 3 contributors of all 3 years
    top_authors = return_top_authors_for_all_years(authors_2022, authors_2023, authors_2024)

    # This function expects all the data of the past 3 years and the top authors
    # This function will return the final dict to place in the excel file
    final_dict = return_author_with_contributions(authors_2022, authors_2023, authors_2024, top_authors)

    # This function expects the final dict and the output file
    # will output the final dict into the excel file without returning anything
    save_to_excel(final_dict, output_file)


# This function takes in all the data we have as dictionaries and the top authors we have
# then it will iterate through the top authors and set their contributions based on the other data we gave this function per year
# Then this will create the final dictionary to place in the excel sheet
def return_author_with_contributions(authors_2022, authors_2023, authors_2024, top_authors):
    top_contributors_data = []
    top_authors_dict = dict(top_authors)
    for author in top_authors:
        contributions_2022 = authors_2022.get(author[0])
        contributions_2023 = authors_2023.get(author[0])
        contributions_2024 = authors_2024.get(author[0])
        top_contributors_data.append({
            'Author': author[0],
            '2022': contributions_2022,
            '2023': contributions_2023,
            '2024': contributions_2024,
            'Total': top_authors_dict[author[0]]
        })
    return top_contributors_data


# This function expects a keyword argument as the url to query
# This function will use get and return the untouched response from the url
def get_most_contributions(**kwargs):
    url = kwargs.get('url')
    return requests.get(url)


# This function will take a raw html response from an API and parse it
# This function is specifically looking for authors and will look for the specific class in the html
# This function relised on bs4 to parse the html
# This function will return a dict of all the authors and how many times they contributed
def parse_authors(**kwargs):
    raw_author_response = kwargs.get('raw_html')
    raw_author_response.encoding = 'utf-8'
    soup = bs4.BeautifulSoup(raw_author_response.content, 'html.parser')
    author_sections = soup.find_all('form', class_='authsearch')
    authors = {}
    for form in author_sections:
        author_name = form.find('a').get_text(strip=True)
        if author_name in authors:
            authors[author_name] += 1
        else:
            authors[author_name] = 1
    return authors


# This cuntion will use the *args to take in any number of arguments as a tuple
# Then this will create a complete dict of all the tuples
# Then this function will sort them and slice the first three off to return the top three
def return_top_authors_for_all_years(*args):
    combined_dict = {}
    for dictionary in args:
        for key, value in dictionary.items():
            if key in combined_dict:
                combined_dict[key] += value
            else:
                combined_dict[key] = value
    sorted_authors = sorted(combined_dict.items(), key=lambda x: x[1], reverse=True)
    return sorted_authors[:3]


# This function takes in arguments for the data and the output file
# Then function will create the data frame, set the author key as the index
# Then it will transpose the data frame (Switch the columns to rows and vice versa)
# Then This will utilize to_excel to print everything to an excel sheet
# This function will also apply borders to everything in the excel sheet
def save_to_excel(data, output_file):
    df = pd.DataFrame(data)
    df.set_index('Author', inplace=True)
    df = df.T
    df.to_excel(output_file)

    workbook = load_workbook(output_file)
    sheet = workbook.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    workbook.save(output_file)



if __name__ == "__main__":
    main()