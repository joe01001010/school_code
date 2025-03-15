import openpyxl
from openpyxl.utils import get_column_letter


def main():
    excel_workbook = "test_workbook.xlsx"
    print(f"openpyxl version: {openpyxl.__version__}")
    print()

    wb = openpyxl.Workbook()
    print(type(wb))
    print()

    new_sheet = wb.create_sheet('New Sheet')
    print(f"New Sheet created: {new_sheet}")
    print()

    sheets = wb.sheetnames
    print(f"All sheets: {sheets}")
    print()

    wb.active = wb.sheetnames.index("New Sheet")
    active_sheet = wb.active
    print(f"Type of active sheet: {type(active_sheet)}")
    print(f"Active sheet: {active_sheet}")
    print()

    for column in range(1,7):
        for row in range(ord('A'), ord('C')):
            cell = f'{chr(row)}{column}'
            active_sheet[cell] = f'Hello from test code {row}'

    print(f"Number of rows in {active_sheet} : {active_sheet.max_row}")
    print(f"Number of columns in {active_sheet}: {active_sheet.max_column}")
    print()

    print(active_sheet['A1'])
    print(active_sheet['A1'].value)
    print()

    a2 = active_sheet['A2']
    print(f"A2 row: {a2.row}")
    print(f"A2 column: {a2.column}")
    print(f"A2 value: {a2.value}")
    print()

    print(f"Column letter for 1: {get_column_letter(1)}")
    print(f"Column letter for 27: {get_column_letter(27)}")
    print(f"Column letter for 900: {get_column_letter(900)}")
    print()

    a1_though_a10 = tuple(active_sheet['A1':'A10'])
    for row in a1_though_a10:
        print(row)
        for cell in row:
            print(cell)
            print(cell.value)
        print()

    wb.save(excel_workbook)


if __name__ == '__main__':
    main()